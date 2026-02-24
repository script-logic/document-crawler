"""
XLSX document parser using openpyxl.
"""

from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from structlog import get_logger

from app.domain.entities import DocumentType

from .base_parser import BaseParser
from .interfaces import ParseError


class XLSXParser(BaseParser):
    """Parser for XLSX spreadsheets."""

    SUPPORTED_EXTENSIONS: ClassVar[set[str]] = {"xlsx", "xlsm", "xltx"}
    DOCUMENT_TYPE: ClassVar[DocumentType] = DocumentType.XLSX

    def _extract_text(self, file_path: Path) -> str:
        """
        Extract text from XLSX using openpyxl.

        Args:
            file_path: Path to XLSX file.

        Returns:
            Extracted text from all cells in all sheets.

        Raises:
            ParseError: If XLSX cannot be read.
        """
        try:
            workbook = load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True,
            )

            text_parts: list[str] = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                sheet_text: list[str] = []
                sheet_text.append(f"--- Sheet: {sheet_name} ---")

                for row in sheet.iter_rows(values_only=True):
                    row_values = [
                        str(cell) for cell in row if cell is not None
                    ]
                    if row_values:
                        sheet_text.append(" | ".join(row_values))

                if len(sheet_text) > 1:
                    text_parts.extend(sheet_text)

            workbook.close()
            return "\n".join(text_parts)

        except InvalidFileException as e:
            raise ParseError(
                "Invalid XLSX file",
                file_path=file_path,
                original_error=e,
            ) from e
        except FileNotFoundError as e:
            raise ParseError("File not found", file_path=file_path) from e
        except PermissionError as e:
            raise ParseError("Permission denied", file_path=file_path) from e
        except Exception as e:
            raise ParseError(
                f"XLSX parsing failed: {e}",
                file_path=file_path,
                original_error=e,
            ) from e

    def extract_metadata(self, file_path: Path) -> dict[str, str | int | None]:
        """
        Extract XLSX metadata.

        Returns:
            Dict with sheet count, etc.
        """
        try:
            workbook = load_workbook(
                filename=file_path,
                read_only=True,
                data_only=True,
            )

            metadata: dict[str, str | int | None] = {
                "sheet_count": len(workbook.sheetnames),
                "sheets": ", ".join(workbook.sheetnames),
            }

            workbook.close()
            return metadata

        except Exception as e:
            logger = get_logger(__name__)
            logger.warning(
                "Failed to extract XLSX metadata",
                file=str(file_path),
                error=str(e),
            )
            return {}
